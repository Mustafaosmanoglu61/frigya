#!/usr/bin/env python3
"""
4-sheet Tax Portfolio Excel builder.
- Renames existing sheets: "Tarihe Göre" → "Tarihe Göre 2026", "Sembole Göre" → "Sembole Göre 2026"
- Adds "Tarihe Göre 2025" and "Sembole Göre 2025" from Vergi Durumu Özeti (TRY)
"""

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
import os, shutil

SRC  = "/sessions/elegant-sharp-turing/mnt/Tax_Portfolilo/2026/2026_Kar_Zarar_Analizi.xlsx"
DEST = "/sessions/elegant-sharp-turing/mnt/Tax_Portfolilo/Kar_Zarar_Analizi_2025_2026.xlsx"

# ── styles ───────────────────────────────────────────────────────────────────
KAR_FILL   = PatternFill("solid", fgColor="E0F0E0")
ZARAR_FILL = PatternFill("solid", fgColor="FFE0E0")
TOT_FILL   = PatternFill("solid", fgColor="D6E4F0")
HDR_FILL   = PatternFill("solid", fgColor="1F4E79")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
ALT_FILL   = PatternFill("solid", fgColor="F5F9FF")

HDR_FONT  = Font(name="Calibri", bold=True,  color="FFFFFF", size=10)
BOLD_FONT = Font(name="Calibri", bold=True,  size=10)
REG_FONT  = Font(name="Calibri",             size=10)
KAR_FONT  = Font(name="Calibri", bold=True,  color="006600", size=10)
ZAR_FONT  = Font(name="Calibri", bold=True,  color="CC0000", size=10)

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

CTR = Alignment(horizontal="center", vertical="center", wrap_text=False)
LFT = Alignment(horizontal="left",   vertical="center")
RGT = Alignment(horizontal="right",  vertical="center")

def set_cell(ws, row, col, value, font=None, fill=None, align=None, border=None, num_fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:    c.font      = font
    if fill:    c.fill      = fill
    if align:   c.alignment = align
    if border:  c.border    = border
    if num_fmt: c.number_format = num_fmt
    return c

# ── 2025 raw data (Vergi Durumu Özeti) ───────────────────────────────────────
# (tarih_str, sembol, adet, kazanc_tl, beyana_tabi_tl)
DATA_2025 = [
    # Örnek — gerçek veri gitignored kalmalı.
    ("2025-06-18", "AAPL",  3.50,   1134.50,    721.98),
    ("2025-07-01", "MSFT", 15.00,  -3485.19,  -3604.12),
]

TL_FMT  = '#,##0.00 ₺'
NUM_FMT = '#,##0.00'
PCT_FMT = '0.00%'
DAT_FMT = 'DD/MM/YYYY'

def auto_width(ws, col_widths):
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

def add_header_row(ws, row, headers, fill=None, font=None):
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font      = font  or HDR_FONT
        c.fill      = fill  or HDR_FILL
        c.alignment = CTR
        c.border    = thin_border()
    ws.row_dimensions[row].height = 22

def style_data_cell(c, font=None, fill=None, align=None, num_fmt=None):
    c.font      = font    or REG_FONT
    c.fill      = fill    or WHITE_FILL
    c.alignment = align   or RGT
    c.border    = thin_border()
    if num_fmt: c.number_format = num_fmt

# ═══════════════════════════════════════════════════════════════════════════════
# BUILD TARIHE GÖRE 2025
# ═══════════════════════════════════════════════════════════════════════════════
def build_tarihe_2025(wb):
    ws = wb.create_sheet("Tarihe Göre 2025")
    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = False

    # ── title ─────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:G1")
    tc = ws["A1"]
    tc.value     = "2025 YILI KÂR / ZARAR ANALİZİ — TARİHE GÖRE  (Kaynak: Midas Vergi Durumu Özeti)"
    tc.font      = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    tc.fill      = HDR_FILL
    tc.alignment = CTR
    ws.row_dimensions[1].height = 24

    # ── header ────────────────────────────────────────────────────────────────
    headers = ["Tarih", "Sembol", "Varlık Adedi", "Kazanç (₺)",
               "Beyana Tabi Kazanç (₺)", "Durum", "ÜFE Farkı (₺)"]
    add_header_row(ws, 2, headers)

    # ── data rows ─────────────────────────────────────────────────────────────
    for ri, (tarih_s, sembol, adet, kazanc, beyana) in enumerate(DATA_2025, start=3):
        alt = (ri % 2 == 0)
        row_fill = ALT_FILL if alt else WHITE_FILL

        is_kar = kazanc > 0
        durum  = "KÂR" if is_kar else "ZARAR"
        d_fill = KAR_FILL   if is_kar else ZARAR_FILL
        d_font = KAR_FONT   if is_kar else ZAR_FONT

        # parse date
        import datetime as dt
        tarih_date = dt.datetime.strptime(tarih_s, "%Y-%m-%d").date()
        ufe_fark = kazanc - beyana  # ÜFE adjustment absorbed into beyana

        cells_data = [
            (tarih_date, REG_FONT, row_fill, CTR, DAT_FMT),
            (sembol,     BOLD_FONT, row_fill, CTR, None),
            (adet,       REG_FONT, row_fill, RGT, '#,##0.####'),
            (kazanc,     REG_FONT, row_fill, RGT, TL_FMT),
            (beyana,     REG_FONT, row_fill, RGT, TL_FMT),
            (durum,      d_font,   d_fill,   CTR, None),
            (ufe_fark,   REG_FONT, row_fill, RGT, TL_FMT),
        ]
        for ci, (val, fnt, fll, aln, nfmt) in enumerate(cells_data, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = fnt; c.fill = fll; c.alignment = aln; c.border = thin_border()
            if nfmt: c.number_format = nfmt

        ws.row_dimensions[ri].height = 16

    # ── totals row ────────────────────────────────────────────────────────────
    n = len(DATA_2025)
    tot_row = 3 + n
    ws.merge_cells(f"A{tot_row}:C{tot_row}")
    tc = ws.cell(row=tot_row, column=1, value=f"TOPLAM ({n} İşlem)")
    tc.font = BOLD_FONT; tc.fill = TOT_FILL; tc.alignment = CTR; tc.border = thin_border()
    ws.cell(row=tot_row, column=2).border = thin_border()
    ws.cell(row=tot_row, column=3).border = thin_border()

    total_kazanc  = sum(r[3] for r in DATA_2025)
    total_beyana  = sum(r[4] for r in DATA_2025)
    total_ufe     = total_kazanc - total_beyana
    n_kar   = sum(1 for r in DATA_2025 if r[3] > 0)
    n_zarar = sum(1 for r in DATA_2025 if r[3] <= 0)

    for ci, (val, nfmt) in enumerate([
        (total_kazanc, TL_FMT),
        (total_beyana, TL_FMT),
    ], start=4):
        c = ws.cell(row=tot_row, column=ci, value=val)
        c.font = BOLD_FONT; c.fill = TOT_FILL; c.alignment = RGT
        c.border = thin_border(); c.number_format = nfmt

    # durum summary
    c = ws.cell(row=tot_row, column=6,
                value=f"KÂR: {n_kar}  |  ZARAR: {n_zarar}")
    c.font = BOLD_FONT; c.fill = TOT_FILL; c.alignment = CTR; c.border = thin_border()

    c = ws.cell(row=tot_row, column=7, value=total_ufe)
    c.font = BOLD_FONT; c.fill = TOT_FILL; c.alignment = RGT
    c.border = thin_border(); c.number_format = TL_FMT

    ws.row_dimensions[tot_row].height = 20

    # ── threshold note ─────────────────────────────────────────────────────────
    note_row = tot_row + 2
    ws.merge_cells(f"A{note_row}:G{note_row}")
    threshold = 18000
    kar_dur = "AŞTI ✓ → Beyan Gerekli" if total_beyana > threshold else "AŞMADI"
    note_val = (f"Beyana Tabi Toplam: ₺{total_beyana:,.2f}  |  "
                f"Eşik: ₺{threshold:,.0f}  |  {kar_dur}")
    nc = ws.cell(row=note_row, column=1, value=note_val)
    color = "006600" if total_beyana > threshold else "CC0000"
    nc.font = Font(name="Calibri", bold=True, color=color, size=10)
    nc.fill = PatternFill("solid", fgColor="FFFBE6")
    nc.alignment = CTR
    nc.border = thin_border()
    ws.row_dimensions[note_row].height = 18

    # ── column widths ──────────────────────────────────────────────────────────
    auto_width(ws, {1:14, 2:10, 3:14, 4:18, 5:22, 6:10, 7:16})

    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# BUILD SEMBOLE GÖRE 2025
# ═══════════════════════════════════════════════════════════════════════════════
def build_sembole_2025(wb):
    ws = wb.create_sheet("Sembole Göre 2025")
    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = False

    # ── aggregate by symbol ────────────────────────────────────────────────────
    sym_data = defaultdict(lambda: {
        "toplam_islem": 0, "basarili": 0, "basarisiz": 0,
        "toplam_adet": 0.0, "net_kazanc": 0.0, "net_beyana": 0.0,
        "toplam_kar": 0.0, "toplam_zarar": 0.0,
        "son_tarih": None,
    })
    import datetime as dt
    for tarih_s, sembol, adet, kazanc, beyana in DATA_2025:
        d   = sym_data[sembol]
        tarih = dt.datetime.strptime(tarih_s, "%Y-%m-%d").date()
        d["toplam_islem"] += 1
        d["toplam_adet"]  += adet
        d["net_kazanc"]   += kazanc
        d["net_beyana"]   += beyana
        if kazanc > 0:
            d["basarili"]     += 1
            d["toplam_kar"]   += kazanc
        else:
            d["basarisiz"]    += 1
            d["toplam_zarar"] += kazanc
        if d["son_tarih"] is None or tarih > d["son_tarih"]:
            d["son_tarih"] = tarih

    # sort by net_kazanc descending
    sorted_syms = sorted(sym_data.items(), key=lambda x: x[1]["net_kazanc"], reverse=True)

    # ── title ──────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:L1")
    tc = ws["A1"]
    tc.value     = "2025 YILI KÂR / ZARAR ANALİZİ — SEMBOLE GÖRE  (Kaynak: Midas Vergi Durumu Özeti)"
    tc.font      = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    tc.fill      = HDR_FILL
    tc.alignment = CTR
    ws.row_dimensions[1].height = 24

    # ── header ─────────────────────────────────────────────────────────────────
    headers = [
        "Sembol", "Son Satış Tarihi",
        "Toplam İşlem", "Başarılı İşlem", "Başarısız İşlem", "Başarı Oranı %",
        "Toplam Adet", "Net Kazanç (₺)", "Net Beyana Tabi (₺)",
        "Kazanç/Zarar %", "Toplam KÂR (₺)", "Toplam ZARAR (₺)",
    ]
    add_header_row(ws, 2, headers)

    # ── data rows ──────────────────────────────────────────────────────────────
    # For K/Z % we use Net Kazanç / Toplam Adet as a proxy (not true % without cost basis)
    # We'll instead show basari_orani as K/Z proxy; mark K/Z % as N/A
    for ri, (sembol, d) in enumerate(sorted_syms, start=3):
        alt = (ri % 2 == 0)
        rf = ALT_FILL if alt else WHITE_FILL
        is_kar = d["net_kazanc"] > 0
        row_fill_net = KAR_FILL if is_kar else ZARAR_FILL
        net_font     = KAR_FONT if is_kar else ZAR_FONT

        basari_oran = d["basarili"] / d["toplam_islem"] if d["toplam_islem"] else 0
        # K/Z % = Net Kazanç / |Toplam Zarar component| - not meaningful without cost basis
        # We'll leave as N/A (None)

        row_vals = [
            (sembol,             BOLD_FONT, rf,            CTR, None),
            (d["son_tarih"],     REG_FONT,  rf,            CTR, DAT_FMT),
            (d["toplam_islem"],  REG_FONT,  rf,            CTR, '0'),
            (d["basarili"],      KAR_FONT,  rf,            CTR, '0'),
            (d["basarisiz"],     ZAR_FONT,  rf,            CTR, '0'),
            (basari_oran,        BOLD_FONT, rf,            CTR, '0.0%'),
            (d["toplam_adet"],   REG_FONT,  rf,            RGT, '#,##0.####'),
            (d["net_kazanc"],    net_font,  row_fill_net,  RGT, TL_FMT),
            (d["net_beyana"],    REG_FONT,  rf,            RGT, TL_FMT),
            ("—",                REG_FONT,  rf,            CTR, None),
            (d["toplam_kar"],    KAR_FONT,  rf,            RGT, TL_FMT),
            (d["toplam_zarar"],  ZAR_FONT,  rf,            RGT, TL_FMT),
        ]
        for ci, (val, fnt, fll, aln, nfmt) in enumerate(row_vals, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = fnt; c.fill = fll; c.alignment = aln; c.border = thin_border()
            if nfmt: c.number_format = nfmt
        ws.row_dimensions[ri].height = 16

    # ── grand totals ────────────────────────────────────────────────────────────
    n_sym   = len(sorted_syms)
    tot_row = 3 + n_sym
    tot_data = sym_data  # already aggregated

    all_islem    = sum(d["toplam_islem"] for d in sym_data.values())
    all_basarili = sum(d["basarili"]     for d in sym_data.values())
    all_basarisiz= sum(d["basarisiz"]    for d in sym_data.values())
    all_adet     = sum(d["toplam_adet"]  for d in sym_data.values())
    all_kazanc   = sum(d["net_kazanc"]   for d in sym_data.values())
    all_beyana   = sum(d["net_beyana"]   for d in sym_data.values())
    all_kar      = sum(d["toplam_kar"]   for d in sym_data.values())
    all_zarar    = sum(d["toplam_zarar"] for d in sym_data.values())
    all_basari_o = all_basarili / all_islem if all_islem else 0

    tot_vals = [
        (f"TOPLAM ({n_sym} Sembol)", BOLD_FONT, TOT_FILL, CTR, None),
        ("—",                         BOLD_FONT, TOT_FILL, CTR, None),
        (all_islem,                   BOLD_FONT, TOT_FILL, CTR, '0'),
        (all_basarili,                BOLD_FONT, TOT_FILL, CTR, '0'),
        (all_basarisiz,               BOLD_FONT, TOT_FILL, CTR, '0'),
        (all_basari_o,                BOLD_FONT, TOT_FILL, CTR, '0.0%'),
        (all_adet,                    BOLD_FONT, TOT_FILL, RGT, '#,##0.####'),
        (all_kazanc,                  BOLD_FONT, TOT_FILL, RGT, TL_FMT),
        (all_beyana,                  BOLD_FONT, TOT_FILL, RGT, TL_FMT),
        ("—",                         BOLD_FONT, TOT_FILL, CTR, None),
        (all_kar,                     BOLD_FONT, TOT_FILL, RGT, TL_FMT),
        (all_zarar,                   BOLD_FONT, TOT_FILL, RGT, TL_FMT),
    ]
    for ci, (val, fnt, fll, aln, nfmt) in enumerate(tot_vals, 1):
        c = ws.cell(row=tot_row, column=ci, value=val)
        c.font = fnt; c.fill = fll; c.alignment = aln; c.border = thin_border()
        if nfmt: c.number_format = nfmt
    ws.row_dimensions[tot_row].height = 20

    # ── threshold note ──────────────────────────────────────────────────────────
    note_row = tot_row + 2
    ws.merge_cells(f"A{note_row}:L{note_row}")
    threshold = 18000
    kar_dur = "AŞTI ✓ → Beyan Gerekli" if all_beyana > threshold else "AŞMADI"
    note_val = (f"Beyana Tabi Toplam: ₺{all_beyana:,.2f}  |  "
                f"Eşik: ₺{threshold:,.0f}  |  {kar_dur}  |  "
                f"Toplam Kazanç: ₺{all_kazanc:,.2f}")
    nc = ws.cell(row=note_row, column=1, value=note_val)
    color = "006600" if all_beyana > threshold else "CC0000"
    nc.font = Font(name="Calibri", bold=True, color=color, size=10)
    nc.fill = PatternFill("solid", fgColor="FFFBE6")
    nc.alignment = CTR
    nc.border = thin_border()
    ws.row_dimensions[note_row].height = 18

    auto_width(ws, {1:10, 2:16, 3:14, 4:14, 5:16, 6:14,
                    7:14, 8:18, 9:20, 10:14, 11:17, 12:17})

    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════
print(f"Loading workbook: {SRC}")
wb = load_workbook(SRC)

# Step 1: rename existing sheets
renamed = []
for ws in wb.worksheets:
    if ws.title == "Tarihe Göre":
        ws.title = "Tarihe Göre 2026"
        renamed.append("Tarihe Göre → Tarihe Göre 2026")
    elif ws.title == "Sembole Göre":
        ws.title = "Sembole Göre 2026"
        renamed.append("Sembole Göre → Sembole Göre 2026")

print("Renamed sheets:", renamed)
print("Current sheets:", [ws.title for ws in wb.worksheets])

# Step 2: build 2025 sheets
print("Building Tarihe Göre 2025 ...")
build_tarihe_2025(wb)

print("Building Sembole Göre 2025 ...")
build_sembole_2025(wb)

# Step 3: reorder sheets — 2026 first, 2025 after
desired_order = [
    "Tarihe Göre 2026",
    "Sembole Göre 2026",
    "Tarihe Göre 2025",
    "Sembole Göre 2025",
]
# openpyxl sheet reorder via _sheets list
name_to_ws = {ws.title: ws for ws in wb.worksheets}
wb._sheets = [name_to_ws[n] for n in desired_order if n in name_to_ws]

# Step 4: save
print(f"Saving to: {DEST}")
wb.save(DEST)
print("Done ✓")
print(f"Total 2025 transactions: {len(DATA_2025)}")
print(f"Unique symbols 2025: {len(set(r[1] for r in DATA_2025))}")
total_k = sum(r[3] for r in DATA_2025)
total_b = sum(r[4] for r in DATA_2025)
print(f"2025 Toplam Kazanç: ₺{total_k:,.2f}")
print(f"2025 Beyana Tabi:   ₺{total_b:,.2f}")
