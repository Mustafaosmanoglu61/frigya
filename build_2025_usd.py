#!/usr/bin/env python3
"""
2025 USD-based FIFO P&L sheets — identical column structure to 2026 sheets.
Source: 2025 monthly PDFs (June–December).
Replaces the TRY-based 2025 sheets in Kar_Zarar_Analizi_2025_2026.xlsx
"""

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
import datetime as dt

import os
XLSX = os.getenv(
    "TAX_XLSX_PATH",
    os.path.join(os.path.dirname(__file__), "Kar_Zarar_Analizi_2025_2026.xlsx"),
)

# ── colour / style constants (identical to 2026 sheets) ──────────────────────
KAR_FILL   = PatternFill("solid", fgColor="E0F0E0")
ZARAR_FILL = PatternFill("solid", fgColor="FFE0E0")
TOT_FILL   = PatternFill("solid", fgColor="D6E4F0")
HDR_FILL   = PatternFill("solid", fgColor="1F4E79")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
ALT_FILL   = PatternFill("solid", fgColor="F5F9FF")
EKS_FILL   = PatternFill("solid", fgColor="FFF3CD")   # amber for "lot missing"

HDR_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
BOLD_FONT  = Font(name="Calibri", bold=True, size=10)
REG_FONT   = Font(name="Calibri", size=10)
KAR_FONT   = Font(name="Calibri", bold=True, color="006600", size=10)
ZAR_FONT   = Font(name="Calibri", bold=True, color="CC0000", size=10)
EKS_FONT   = Font(name="Calibri", bold=True, color="856404", size=10)

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

CTR = Alignment(horizontal="center", vertical="center")
LFT = Alignment(horizontal="left",   vertical="center")
RGT = Alignment(horizontal="right",  vertical="center")

USD_FMT  = '#,##0.00 $'
NUM_FMT  = '#,##0.####'
PCT_FMT  = '0.00%'
DAT_FMT  = 'DD/MM/YYYY'

def auto_width(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

def hdr_row(ws, row, headers):
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = HDR_FONT; c.fill = HDR_FILL
        c.alignment = CTR; c.border = thin_border()
    ws.row_dimensions[row].height = 22

# ═══════════════════════════════════════════════════════════════════════════════
# 2025 TRANSACTIONS
# Real data lives in data/transactions_2025.py (gitignored).
# Copy data/transactions_2025.sample.py → data/transactions_2025.py and fill in.
# ═══════════════════════════════════════════════════════════════════════════════
def _load_tx():
    import importlib.util, pathlib
    base = pathlib.Path(__file__).parent / "data"
    real = base / "transactions_2025.py"
    sample = base / "transactions_2025.sample.py"
    target = real if real.exists() else sample
    spec = importlib.util.spec_from_file_location("_tx_2025", target)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod.TX

TX = _load_tx()


# ═══════════════════════════════════════════════════════════════════════════════
# FIFO ENGINE
# ═══════════════════════════════════════════════════════════════════════════════
fifo = defaultdict(list)   # symbol → list of (qty, price, cost)

sell_results = []   # each item: dict with all fields for the Excel

for tarih_s, sembol, islem, adet, fiyat, toplam in TX:
    tarih = dt.datetime.strptime(tarih_s, "%Y-%m-%d").date()

    if islem == "Alış":
        fifo[sembol].append([adet, fiyat, toplam])   # [qty, price, cost]

    else:  # Satış
        satis_geliri   = toplam
        kalan          = adet
        total_cost     = 0.0
        eksik_lot      = False

        while kalan > 1e-7:
            if not fifo[sembol]:
                eksik_lot = True
                # cost for remaining unknown → leave as 0
                break

            lot = fifo[sembol][0]
            lot_qty, lot_price, lot_cost = lot

            if lot_qty <= kalan + 1e-7:
                consumed   = min(lot_qty, kalan)
                frac       = consumed / lot_qty if lot_qty > 1e-10 else 0
                total_cost += lot_cost * frac
                kalan      -= consumed
                fifo[sembol].pop(0)
            else:
                oran       = kalan / lot_qty
                partial    = lot_cost * oran
                total_cost += partial
                fifo[sembol][0] = [lot_qty - kalan, lot_price, lot_cost - partial]
                kalan = 0

        kar_zarar = satis_geliri - total_cost

        sell_results.append({
            "tarih":         tarih,
            "sembol":        sembol,
            "adet":          adet,
            "satis_fiyati":  fiyat,
            "satis_geliri":  satis_geliri,
            "alis_maliyeti": total_cost,
            "kar_zarar":     kar_zarar,
            "eksik_lot":     eksik_lot,
        })

print(f"Total sell transactions: {len(sell_results)}")
n_eksik = sum(1 for r in sell_results if r["eksik_lot"])
print(f"Incomplete lots (eksik): {n_eksik}")
for r in sell_results:
    if r["eksik_lot"]:
        print(f"  EKSIK: {r['tarih']} {r['sembol']} {r['adet']:.4f}")

total_gelir = sum(r["satis_geliri"]  for r in sell_results)
total_maliyet = sum(r["alis_maliyeti"] for r in sell_results)
total_kz = sum(r["kar_zarar"] for r in sell_results)
print(f"Total satış geliri:  ${total_gelir:,.2f}")
print(f"Total alış maliyeti: ${total_maliyet:,.2f}")
print(f"Total K/Z:           ${total_kz:,.2f}")


# ═══════════════════════════════════════════════════════════════════════════════
# OPEN EXISTING WORKBOOK & REPLACE 2025 SHEETS
# ═══════════════════════════════════════════════════════════════════════════════
wb = load_workbook(XLSX)
print("Existing sheets:", [ws.title for ws in wb.worksheets])

# Remove old 2025 sheets if present
for old_name in ["Tarihe Göre 2025", "Sembole Göre 2025"]:
    if old_name in wb.sheetnames:
        del wb[old_name]
        print(f"  Removed: {old_name}")


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — TARİHE GÖRE 2025 (same columns as Tarihe Göre 2026)
# ═══════════════════════════════════════════════════════════════════════════════
ws_t = wb.create_sheet("Tarihe Göre 2025")
ws_t.freeze_panes = "A3"
ws_t.sheet_view.showGridLines = False

# title
ws_t.merge_cells("A1:I1")
tc = ws_t["A1"]
tc.value     = "2025 YILI KÂR / ZARAR ANALİZİ — TARİHE GÖRE  (USD, FIFO, Kaynak: Aylık Ekstreler)"
tc.font      = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
tc.fill      = HDR_FILL
tc.alignment = CTR
ws_t.row_dimensions[1].height = 24

# headers — identical to 2026
headers_t = [
    "Tarih", "Sembol", "Satış Adedi", "Satış Fiyatı (USD)",
    "Satış Geliri (USD)", "Alış Maliyeti (USD)",
    "Kâr / Zarar (USD)", "Durum", "Kâr/Zarar\nYüzdesi %"
]
hdr_row(ws_t, 2, headers_t)

for ri, r in enumerate(sell_results, start=3):
    alt = (ri % 2 == 0)
    rf  = ALT_FILL if alt else WHITE_FILL

    is_kar = r["kar_zarar"] >= 0
    if r["eksik_lot"]:
        durum      = "KÂR*" if is_kar else "ZAR*"
        d_fill     = EKS_FILL
        d_font     = EKS_FONT
    else:
        durum      = "KÂR" if is_kar else "ZARAR"
        d_fill     = KAR_FILL   if is_kar else ZARAR_FILL
        d_font     = KAR_FONT   if is_kar else ZAR_FONT

    pct = (r["kar_zarar"] / r["alis_maliyeti"]) if r["alis_maliyeti"] > 0.001 else None

    row_data = [
        (r["tarih"],         REG_FONT,  rf,     CTR, DAT_FMT),
        (r["sembol"],        BOLD_FONT, rf,     CTR, None),
        (r["adet"],          REG_FONT,  rf,     RGT, '#,##0.####'),
        (r["satis_fiyati"],  REG_FONT,  rf,     RGT, '#,##0.00 $'),
        (r["satis_geliri"],  REG_FONT,  rf,     RGT, USD_FMT),
        (r["alis_maliyeti"], REG_FONT,  rf,     RGT, USD_FMT),
        (r["kar_zarar"],     REG_FONT,  rf,     RGT, USD_FMT),
        (durum,              d_font,    d_fill, CTR, None),
        (pct,                REG_FONT,  rf,     RGT, '0.00%'),
    ]
    for ci, (val, fnt, fll, aln, nfmt) in enumerate(row_data, 1):
        c = ws_t.cell(row=ri, column=ci, value=val)
        c.font = fnt; c.fill = fll; c.alignment = aln; c.border = thin_border()
        if nfmt and val is not None: c.number_format = nfmt
    ws_t.row_dimensions[ri].height = 16

# totals row
n = len(sell_results)
tot_r = 3 + n
ws_t.merge_cells(f"A{tot_r}:C{tot_r}")
c = ws_t.cell(row=tot_r, column=1, value=f"TOPLAM ({n} İşlem)")
c.font = BOLD_FONT; c.fill = TOT_FILL; c.alignment = CTR; c.border = thin_border()
ws_t.cell(row=tot_r,column=2).border = thin_border()
ws_t.cell(row=tot_r,column=3).border = thin_border()

n_kar   = sum(1 for r in sell_results if r["kar_zarar"] >= 0)
n_zarar = sum(1 for r in sell_results if r["kar_zarar"] <  0)
n_eksik = sum(1 for r in sell_results if r["eksik_lot"])

for ci, (val, nfmt) in enumerate([
    (total_gelir,    USD_FMT),
    (total_maliyet,  USD_FMT),
    (total_kz,       USD_FMT),
], start=5):
    c = ws_t.cell(row=tot_r, column=ci, value=val)
    c.font = BOLD_FONT; c.fill = TOT_FILL; c.alignment = RGT
    c.border = thin_border(); c.number_format = nfmt

c = ws_t.cell(row=tot_r, column=4, value="")
c.fill = TOT_FILL; c.border = thin_border()

c = ws_t.cell(row=tot_r, column=8, value=f"KÂR: {n_kar}  ZARAR: {n_zarar}")
c.font = BOLD_FONT; c.fill = TOT_FILL; c.alignment = CTR; c.border = thin_border()

pct_tot = (total_kz / total_maliyet) if total_maliyet > 0.001 else None
c = ws_t.cell(row=tot_r, column=9, value=pct_tot)
c.font = BOLD_FONT; c.fill = TOT_FILL; c.alignment = RGT
c.border = thin_border()
if pct_tot is not None: c.number_format = '0.00%'

ws_t.row_dimensions[tot_r].height = 20

# footnote
if n_eksik > 0:
    note_row = tot_r + 2
    ws_t.merge_cells(f"A{note_row}:I{note_row}")
    nc = ws_t.cell(row=note_row, column=1,
        value=f"* {n_eksik} işlemde önceki dönemden devir lot bulunmadığı için alış maliyeti eksik hesaplanmıştır. "
              f"Durum kolonu 'KÂR*' / 'ZAR*' ile işaretlendi.")
    nc.font = EKS_FONT
    nc.fill = PatternFill("solid", fgColor="FFFBE6")
    nc.alignment = LFT; nc.border = thin_border()
    ws_t.row_dimensions[note_row].height = 18

auto_width(ws_t, {1:14, 2:12, 3:14, 4:18, 5:18, 6:18, 7:16, 8:9, 9:14})


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — SEMBOLE GÖRE 2025 (same columns as Sembole Göre 2026)
# ═══════════════════════════════════════════════════════════════════════════════
sym = defaultdict(lambda: {
    "toplam_islem": 0, "basarili": 0, "basarisiz": 0,
    "toplam_adet": 0.0,
    "satis_geliri": 0.0, "alis_maliyeti": 0.0, "net_kz": 0.0,
    "toplam_kar": 0.0, "toplam_zarar": 0.0,
    "son_tarih": None, "son_fiyat": 0.0,
})

for r in sell_results:
    d = sym[r["sembol"]]
    d["toplam_islem"]  += 1
    d["toplam_adet"]   += r["adet"]
    d["satis_geliri"]  += r["satis_geliri"]
    d["alis_maliyeti"] += r["alis_maliyeti"]
    d["net_kz"]        += r["kar_zarar"]
    if r["kar_zarar"] >= 0:
        d["basarili"]    += 1
        d["toplam_kar"]  += r["kar_zarar"]
    else:
        d["basarisiz"]   += 1
        d["toplam_zarar"] += r["kar_zarar"]
    if d["son_tarih"] is None or r["tarih"] > d["son_tarih"]:
        d["son_tarih"]  = r["tarih"]
        d["son_fiyat"]  = r["satis_fiyati"]

# sort by net_kz descending
sorted_syms = sorted(sym.items(), key=lambda x: x[1]["net_kz"], reverse=True)

ws_s = wb.create_sheet("Sembole Göre 2025")
ws_s.freeze_panes = "A3"
ws_s.sheet_view.showGridLines = False

# title
ws_s.merge_cells("A1:N1")
tc = ws_s["A1"]
tc.value     = "2025 YILI KÂR / ZARAR ANALİZİ — SEMBOLE GÖRE  (USD, FIFO, Kaynak: Aylık Ekstreler)"
tc.font      = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
tc.fill      = HDR_FILL
tc.alignment = CTR
ws_s.row_dimensions[1].height = 24

# headers — identical to Sembole Göre 2026
headers_s = [
    "Sembol", "Son Satış\nTarihi", "Son Satış\nFiyatı (USD)",
    "Toplam\nİşlem", "Başarılı\nİşlem", "Başarısız\nİşlem",
    "Başarı\nOranı %", "Toplam Adet",
    "Satış Geliri\n(USD)", "Alış Maliyeti\n(USD)",
    "Net Kâr/Zarar\n(USD)", "Kâr/Zarar\nYüzdesi %",
    "Toplam KÂR\n(USD)", "Toplam ZARAR\n(USD)"
]
hdr_row(ws_s, 2, headers_s)

for ri, (sembol, d) in enumerate(sorted_syms, start=3):
    alt = (ri % 2 == 0)
    rf  = ALT_FILL if alt else WHITE_FILL

    is_kar   = d["net_kz"] >= 0
    net_fill = KAR_FILL if is_kar else ZARAR_FILL
    net_font = KAR_FONT if is_kar else ZAR_FONT
    bas_oran = d["basarili"] / d["toplam_islem"] if d["toplam_islem"] else 0
    pct      = (d["net_kz"] / d["alis_maliyeti"]) if d["alis_maliyeti"] > 0.001 else None

    row_data = [
        (sembol,              BOLD_FONT, rf,       CTR, None),
        (d["son_tarih"],      REG_FONT,  rf,       CTR, DAT_FMT),
        (d["son_fiyat"],      REG_FONT,  rf,       RGT, '#,##0.00 $'),
        (d["toplam_islem"],   REG_FONT,  rf,       CTR, '0'),
        (d["basarili"],       KAR_FONT,  rf,       CTR, '0'),
        (d["basarisiz"],      ZAR_FONT,  rf,       CTR, '0'),
        (bas_oran,            BOLD_FONT, rf,       CTR, '0.0%'),
        (d["toplam_adet"],    REG_FONT,  rf,       RGT, '#,##0.####'),
        (d["satis_geliri"],   REG_FONT,  rf,       RGT, USD_FMT),
        (d["alis_maliyeti"],  REG_FONT,  rf,       RGT, USD_FMT),
        (d["net_kz"],         net_font,  net_fill, RGT, USD_FMT),
        (pct,                 REG_FONT,  rf,       RGT, '0.00%'),
        (d["toplam_kar"],     KAR_FONT,  rf,       RGT, USD_FMT),
        (d["toplam_zarar"],   ZAR_FONT,  rf,       RGT, USD_FMT),
    ]
    for ci, (val, fnt, fll, aln, nfmt) in enumerate(row_data, 1):
        c = ws_s.cell(row=ri, column=ci, value=val)
        c.font = fnt; c.fill = fll; c.alignment = aln; c.border = thin_border()
        if nfmt and val is not None: c.number_format = nfmt
    ws_s.row_dimensions[ri].height = 16

# grand totals
n_sym   = len(sorted_syms)
tot_r   = 3 + n_sym
all_islem    = sum(d["toplam_islem"]  for _,d in sorted_syms)
all_bas      = sum(d["basarili"]      for _,d in sorted_syms)
all_bas_n    = sum(d["basarisiz"]     for _,d in sorted_syms)
all_adet     = sum(d["toplam_adet"]   for _,d in sorted_syms)
all_gelir    = sum(d["satis_geliri"]  for _,d in sorted_syms)
all_maliyet  = sum(d["alis_maliyeti"] for _,d in sorted_syms)
all_kz       = sum(d["net_kz"]        for _,d in sorted_syms)
all_kar      = sum(d["toplam_kar"]    for _,d in sorted_syms)
all_zarar    = sum(d["toplam_zarar"]  for _,d in sorted_syms)
all_bas_oran = all_bas / all_islem if all_islem else 0
all_pct      = (all_kz / all_maliyet) if all_maliyet > 0.001 else None

tot_vals = [
    (f"TOPLAM ({n_sym} Sembol)", BOLD_FONT, TOT_FILL, CTR, None),
    ("—",         BOLD_FONT, TOT_FILL, CTR, None),
    ("—",         BOLD_FONT, TOT_FILL, CTR, None),
    (all_islem,   BOLD_FONT, TOT_FILL, CTR, '0'),
    (all_bas,     BOLD_FONT, TOT_FILL, CTR, '0'),
    (all_bas_n,   BOLD_FONT, TOT_FILL, CTR, '0'),
    (all_bas_oran,BOLD_FONT, TOT_FILL, CTR, '0.0%'),
    (all_adet,    BOLD_FONT, TOT_FILL, RGT, '#,##0.####'),
    (all_gelir,   BOLD_FONT, TOT_FILL, RGT, USD_FMT),
    (all_maliyet, BOLD_FONT, TOT_FILL, RGT, USD_FMT),
    (all_kz,      BOLD_FONT, TOT_FILL, RGT, USD_FMT),
    (all_pct,     BOLD_FONT, TOT_FILL, RGT, '0.00%'),
    (all_kar,     BOLD_FONT, TOT_FILL, RGT, USD_FMT),
    (all_zarar,   BOLD_FONT, TOT_FILL, RGT, USD_FMT),
]
for ci, (val, fnt, fll, aln, nfmt) in enumerate(tot_vals, 1):
    c = ws_s.cell(row=tot_r, column=ci, value=val)
    c.font = fnt; c.fill = fll; c.alignment = aln; c.border = thin_border()
    if nfmt and val is not None: c.number_format = nfmt
ws_s.row_dimensions[tot_r].height = 20

auto_width(ws_s, {1:14, 2:14, 3:16, 4:10, 5:10, 6:12, 7:10,
                  8:14, 9:16, 10:17, 11:16, 12:13, 13:15, 14:15})

# ── reorder sheets ────────────────────────────────────────────────────────────
desired = ["Tarihe Göre 2026","Sembole Göre 2026",
           "Tarihe Göre 2025","Sembole Göre 2025"]
name_map = {ws.title: ws for ws in wb.worksheets}
wb._sheets = [name_map[n] for n in desired if n in name_map]

# ── save ──────────────────────────────────────────────────────────────────────
wb.save(XLSX)
print(f"\nSaved: {XLSX}")
print("Sheets:", [ws.title for ws in wb.worksheets])
print(f"\n2025 USD Summary:")
print(f"  Sell transactions : {len(sell_results)}")
print(f"  Unique symbols    : {n_sym}")
print(f"  Total proceeds    : ${all_gelir:,.2f}")
print(f"  Total cost        : ${all_maliyet:,.2f}")
print(f"  Net K/Z           : ${all_kz:,.2f}")
print(f"  Başarı oranı      : {all_bas_oran:.1%}")
print(f"  Eksik lot count   : {n_eksik}")
